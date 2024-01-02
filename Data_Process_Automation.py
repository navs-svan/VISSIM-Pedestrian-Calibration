import openpyxl as op

filename = r'D:\Drive D\ACADS\THESIS\Data\Validation 1\Sequence1.xlsx'
wb_read = op.load_workbook(filename, data_only=True)
ws_read = wb_read['Track2']

wb_write = op.load_workbook(filename)
ws_write = wb_write['Track2']

max_row = ws_write.max_row
num_last_person = ws_write[f'B{max_row}'].value

i = 1
first_num = 3
last_num = 2

while i <= max_row:
    i += 1
    if ws_read[f'B{i}'].value != ws_read[f'B{i + 1}'].value:
        last_num = i
        ws_write[f'K{last_num}'] = f'=SUM(J{first_num}:J{last_num})/C{last_num}'
        first_num = last_num + 2

wb_write.save(r'D:\Drive D\ACADS\THESIS\Data\Validation 1\Sequence1_modified.xlsx')
